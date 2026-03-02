"""
XLSX Template Generator for Livestock Register

Creates the Google Sheets-compatible template with all 6 sheets
and proper column headers, formatting, and data validation.

Usage:
    python create_template.py

Output:
    ../data/livestock_template.xlsx
"""

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


def create_header_style() -> tuple[Font, PatternFill, Alignment, Border]:
    """Create consistent header styling for all sheets."""
    font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    return font, fill, alignment, border


def apply_headers(ws, headers: list[str], widths: list[int]) -> None:
    """Apply headers with styling and column widths to a worksheet."""
    font, fill, alignment, border = create_header_style()

    for col_idx, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = font
        cell.fill = fill
        cell.alignment = alignment
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.freeze_panes = "A2"


def create_registro_sheet(wb: openpyxl.Workbook) -> None:
    """Create the 'Registro' sheet for animal records."""
    ws = wb.active
    ws.title = "Registro"

    headers = [
        "animal_id",
        "arete_id",
        "nombre",
        "tipo",
        "sexo",
        "fecha_nacimiento",
        "raza",
        "madre_id",
        "padre_id",
        "temperamento",
        "estado",
        "peso_actual",
        "notas",
        "foto_url",
        "created_by",
        "updated_at",
        "created_at",
    ]
    widths = [12, 15, 15, 14, 10, 18, 15, 12, 12, 14, 14, 18, 30, 20, 12, 20, 20]

    apply_headers(ws, headers, widths)

    # Data validations
    tipo_dv = DataValidation(
        type="list",
        formula1='"Semental,Becerro(a),Vaquilla,Vaca,Torete"',
        allow_blank=True,
    )
    tipo_dv.error = "Seleccione un tipo válido"
    tipo_dv.errorTitle = "Tipo inválido"
    ws.add_data_validation(tipo_dv)
    tipo_dv.add(f"D2:D1000")

    sexo_dv = DataValidation(
        type="list", formula1='"Macho,Hembra"', allow_blank=True
    )
    ws.add_data_validation(sexo_dv)
    sexo_dv.add(f"E2:E1000")

    temperamento_dv = DataValidation(
        type="list",
        formula1='"Normal,Manso(a),Bravo(a)"',
        allow_blank=True,
    )
    ws.add_data_validation(temperamento_dv)
    temperamento_dv.add(f"J2:J1000")

    estado_dv = DataValidation(
        type="list",
        formula1='"Vivo(a),Muerto(a),Vendido(a)"',
        allow_blank=True,
    )
    ws.add_data_validation(estado_dv)
    estado_dv.add(f"K2:K1000")


def create_salud_sheet(wb: openpyxl.Workbook) -> None:
    """Create the 'Salud' sheet for health records."""
    ws = wb.create_sheet("Salud")

    headers = [
        "salud_id",
        "animal_id",
        "fecha",
        "tipo_evento",
        "producto",
        "dosis",
        "estado_general",
        "proxima_aplicacion",
        "notas",
        "created_by",
        "updated_at",
        "created_at",
    ]
    widths = [12, 12, 14, 18, 20, 12, 16, 20, 30, 12, 20, 20]

    apply_headers(ws, headers, widths)

    tipo_evento_dv = DataValidation(
        type="list",
        formula1='"Vacuna,Desparasitación,Vitamina,Enfermedad,Tratamiento,Revisión"',
        allow_blank=True,
    )
    ws.add_data_validation(tipo_evento_dv)
    tipo_evento_dv.add(f"D2:D1000")

    estado_dv = DataValidation(
        type="list",
        formula1='"Fuerte,Flaco,Enfermo"',
        allow_blank=True,
    )
    ws.add_data_validation(estado_dv)
    estado_dv.add(f"G2:G1000")


def create_reproduccion_sheet(wb: openpyxl.Workbook) -> None:
    """Create the 'Reproduccion' sheet for breeding records."""
    ws = wb.create_sheet("Reproduccion")

    headers = [
        "reproduccion_id",
        "vaca_id",
        "semental_id",
        "fecha_monta",
        "fecha_posible_parto",
        "prenez_confirmada",
        "fecha_parto_real",
        "cria_id",
        "peso_destete_cria",
        "notas",
        "created_by",
        "updated_at",
        "created_at",
    ]
    widths = [18, 12, 14, 16, 20, 18, 18, 12, 22, 30, 12, 20, 20]

    apply_headers(ws, headers, widths)

    prenez_dv = DataValidation(
        type="list",
        formula1='"Sí,No,Pendiente"',
        allow_blank=True,
    )
    ws.add_data_validation(prenez_dv)
    prenez_dv.add(f"F2:F1000")


def create_observaciones_sheet(wb: openpyxl.Workbook) -> None:
    """Create the 'Observaciones' sheet for field observations."""
    ws = wb.create_sheet("Observaciones")

    headers = [
        "observacion_id",
        "fecha",
        "animal_id",
        "notas",
        "created_by",
        "updated_at",
        "created_at",
    ]
    widths = [18, 14, 12, 40, 12, 20, 20]

    apply_headers(ws, headers, widths)


def create_ventas_sheet(wb: openpyxl.Workbook) -> None:
    """Create the 'Ventas' sheet for sales records."""
    ws = wb.create_sheet("Ventas")

    headers = [
        "venta_id",
        "animal_id",
        "fecha_venta",
        "motivo_venta",
        "peso",
        "precio_total",
        "precio_kg",
        "comprador",
        "notas",
        "created_by",
        "updated_at",
        "created_at",
    ]
    widths = [12, 12, 14, 20, 12, 16, 18, 18, 30, 12, 20, 20]

    apply_headers(ws, headers, widths)

    motivo_dv = DataValidation(
        type="list",
        formula1='"Por peso (destete),Por edad,Por productividad,Otro"',
        allow_blank=True,
    )
    ws.add_data_validation(motivo_dv)
    motivo_dv.add(f"D2:D1000")


def create_recorridos_sheet(wb: openpyxl.Workbook) -> None:
    """Create the 'Recorridos' sheet for field patrol records."""
    ws = wb.create_sheet("Recorridos")

    headers = [
        "recorrido_id",
        "fecha",
        "animal_id",
        "notas",
        "created_by",
        "updated_at",
        "created_at",
    ]
    widths = [18, 14, 12, 40, 12, 20, 20]

    apply_headers(ws, headers, widths)


def create_usuarios_sheet(wb: openpyxl.Workbook) -> None:
    """Create the 'Usuarios' sheet for user management."""
    ws = wb.create_sheet("Usuarios")

    headers = [
        "user_id",
        "nombre",
        "pin_hash",
        "created_at",
    ]
    widths = [12, 20, 40, 20]

    apply_headers(ws, headers, widths)


def main() -> None:
    """Create the livestock register XLSX template."""
    output_dir = Path(__file__).parent.parent / "data"
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / "livestock_template.xlsx"

    wb = openpyxl.Workbook()

    create_registro_sheet(wb)
    create_salud_sheet(wb)
    create_reproduccion_sheet(wb)
    create_observaciones_sheet(wb)
    create_ventas_sheet(wb)
    create_recorridos_sheet(wb)
    create_usuarios_sheet(wb)

    wb.save(output_path)
    print(f"Template created: {output_path}")


if __name__ == "__main__":
    main()
